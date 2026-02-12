"""
Outreach Drafter - Standalone script for drafting LinkedIn outreach messages.
Reads applied (status=done) companies from Excel tracker, generates template-based
drafts per company. Uses stable filenames with skip logic so re-runs don't
regenerate drafts unless roles change.
"""

import os
import re
import openpyxl
import shutil
import tempfile
from datetime import datetime
from collections import OrderedDict

from config import TRACKER_FILE, LINKEDIN_URL, USER_PROFILE
from daily_job_search import parse_hr_contacts

OUTPUT_DIR = r'C:\Users\mahas\OneDrive\Desktop\Applications\JobSearch\output'

# --- Profile values from config ---
EXPERIENCE_YEARS = USER_PROFILE.get('experience_years', 5)
LOCATION = USER_PROFILE.get('location', 'City, Country')
_BACKGROUND = USER_PROFILE.get('background', 'a software engineer')
_DOMAIN = USER_PROFILE.get('domain_expertise', 'software development')
_NAME = USER_PROFILE.get('name', 'Your Name')
_ORIGIN = USER_PROFILE.get('origin_country', '')

# --- Templates (all personal info comes from USER_PROFILE in config.py) ---

SHORT_SINGLE_ROLE = (
    "Hi {first_name}, I applied for the {role} role at {company}.\n"
    "With {years}+ yrs in {domain}, I believe I'm a strong fit.\n"
    "Could you check on my application or direct me to the right person?\n"
    "My Professional background:{linkedin}\n"
    "Thanks & regards, {user_name}"
)

SHORT_MULTI_ROLE_ALL = (
    "Hi {first_name}, I applied for {roles_csv} at {company}.\n"
    "With {years}+ yrs in {domain}, I believe I'm a strong fit.\n"
    "Could you check on my application or direct me to the right person?\n"
    "My Professional background:{linkedin}\n"
    "Thanks & regards, {user_name}"
)

SHORT_MULTI_ROLE_SUMMARY = (
    "Hi {first_name}, I applied for {count} roles at {company} "
    "including {first_role}.\n"
    "With {years}+ yrs in {domain}, could you check on my application "
    "or direct me to the right person?\n"
    "My Professional background:{linkedin}\n"
    "Thanks & regards, {user_name}"
)

LONG_TEMPLATE = (
    "Hi [Name],\n\n"
    "My name is {user_name}. I am from {origin_country}, currently based in {location}.\n\n"
    "{{role_section}}\n\n"
    "With {years}+ years of experience in {domain}, I am confident my profile "
    "aligns well with what {{company}} is looking for.\n\n"
    "You can find my professional background here: {linkedin}\n\n"
    "I would be grateful if you could kindly check on the status of my application. "
    "If this does not fall under your purview, I would appreciate being redirected "
    "to the right person who handles these roles.\n\n"
    "Thank you for your time and consideration.\n\n"
    "Best regards,\n"
    "{user_name}"
).format(
    user_name=_NAME,
    origin_country=_ORIGIN,
    location=LOCATION,
    years=EXPERIENCE_YEARS,
    domain=_DOMAIN,
    linkedin=LINKEDIN_URL,
)


def _safe_company_name(company):
    """Convert company name to a filesystem-safe string."""
    return re.sub(r'[^a-z0-9_]', '_', company.lower().strip()).strip('_')


def read_applied_companies():
    """Read Excel tracker, return all status=done companies with ALL their roles and contacts.
    Returns OrderedDict: {company: {'roles': [str], 'contacts': [(name, url)]}}"""
    temp_file = None
    try:
        try:
            wb = openpyxl.load_workbook(TRACKER_FILE, data_only=False)
        except PermissionError:
            print("  File locked - reading from temp copy...")
            fd, temp_file = tempfile.mkstemp(suffix='.xlsx')
            os.close(fd)
            shutil.copy2(TRACKER_FILE, temp_file)
            wb = openpyxl.load_workbook(temp_file, data_only=False)

        ws = wb.active
        applied = OrderedDict()

        for i, row in enumerate(ws.iter_rows()):
            if i == 0:
                continue
            company = str(row[0].value or '').strip()
            role = str(row[1].value or '').strip()
            status = str(row[3].value or '').strip()

            if not company or 'Program/Product' in company:
                continue
            if 'done' not in status.lower():
                continue

            if company not in applied:
                applied[company] = {'roles': [], 'contacts': set()}

            if role and role != 'None' and not role.startswith('http'):
                if role not in applied[company]['roles']:
                    applied[company]['roles'].append(role)

            hr = parse_hr_contacts(row[4] if len(row) > 4 else None)
            for name, url in hr:
                applied[company]['contacts'].add((name, url))

        wb.close()
        if temp_file and os.path.exists(temp_file):
            os.remove(temp_file)

        # Default roles for companies with none listed
        for data in applied.values():
            if not data['roles']:
                data['roles'] = ['Open Positions']

        print(f"  Applied companies found: {len(applied)}")
        return applied
    except Exception as e:
        if temp_file and os.path.exists(temp_file):
            os.remove(temp_file)
        print(f"  Warning: Could not read tracker: {e}")
        return OrderedDict()


def has_existing_draft(company, roles):
    """Check if a draft file already exists with matching roles.
    Returns True if draft is up to date, False if regeneration needed."""
    safe_name = _safe_company_name(company)
    filepath = os.path.join(OUTPUT_DIR, f'outreach_drafts_{safe_name}.txt')

    if not os.path.exists(filepath):
        return False

    try:
        with open(filepath, 'r', encoding='utf-8') as f:
            content = f.read(2000)  # Read just the header portion

        # Extract roles from the header
        # Single role format: "Role: <role>"
        # Multi role format: "Roles Applied (N):\n  - role1\n  - role2"
        existing_roles = []

        single_match = re.search(r'^Role: (.+)$', content, re.MULTILINE)
        if single_match:
            existing_roles = [single_match.group(1).strip()]
        else:
            multi_match = re.findall(r'^\s+-\s+(.+)$', content, re.MULTILINE)
            if multi_match:
                existing_roles = [r.strip() for r in multi_match]

        # Compare sorted role lists
        if sorted(existing_roles) == sorted(roles):
            return True

        return False
    except Exception:
        return False


def _draft_short_message(company, roles, contact_name):
    """Generate a short LinkedIn connection request (max 300 chars)."""
    first_name = contact_name.split()[0].title()
    fmt = {
        'first_name': first_name,
        'company': company,
        'years': EXPERIENCE_YEARS,
        'background': _BACKGROUND,
        'domain': _DOMAIN,
        'linkedin': LINKEDIN_URL,
        'user_name': _NAME,
    }

    if len(roles) == 1:
        msg = SHORT_SINGLE_ROLE.format(role=roles[0], **fmt)
    else:
        roles_csv = ", ".join(roles)
        msg = SHORT_MULTI_ROLE_ALL.format(roles_csv=roles_csv, **fmt)
        if len(msg) > 300:
            msg = SHORT_MULTI_ROLE_SUMMARY.format(
                count=len(roles), first_role=roles[0], **fmt
            )
    return msg


def _draft_long_message(company, roles):
    """Generate a longer InMail/email message."""
    if len(roles) == 1:
        role_section = f"I recently applied for the {roles[0]} role at {company}."
    else:
        roles_csv = ", ".join(roles)
        role_section = (
            f"I recently applied for the following roles at {company}: "
            f"{roles_csv}."
        )

    return LONG_TEMPLATE.format(role_section=role_section, company=company)


def generate_outreach(company, roles, contacts):
    """Generate and save outreach drafts for one company.
    Returns the filepath of the saved draft."""
    safe_name = _safe_company_name(company)
    filepath = os.path.join(OUTPUT_DIR, f'outreach_drafts_{safe_name}.txt')

    contacts_sorted = sorted(contacts, key=lambda x: x[0])
    drafts = []

    # Short LinkedIn message for each contact
    for name, url in contacts_sorted:
        msg = _draft_short_message(company, roles, name)
        drafts.append({
            'contact': name,
            'contact_url': url,
            'type': 'LinkedIn Connection Request (max 300 chars)',
            'message': msg,
        })

    # Long InMail/email template
    long_msg = _draft_long_message(company, roles)
    drafts.append({
        'contact': 'Any HR Contact (InMail / Email)',
        'contact_url': '',
        'type': 'LinkedIn InMail or Email (longer format)',
        'message': long_msg,
    })

    # Write the file
    lines = []
    lines.append(
        f"Outreach Drafts for {company} - "
        f"{datetime.now().strftime('%Y-%m-%d %H:%M')}\n"
    )
    lines.append('=' * 60 + '\n')
    lines.append(f"Company: {company}\n")
    if len(roles) == 1:
        lines.append(f"Role: {roles[0]}\n")
    else:
        lines.append(f"Roles Applied ({len(roles)}):\n")
        for r in roles:
            lines.append(f"  - {r}\n")
    lines.append('=' * 60 + '\n\n')

    for i, draft in enumerate(drafts, 1):
        lines.append(f"--- Draft {i}: {draft['contact']} ---\n")
        lines.append(f"Type: {draft['type']}\n")
        if draft['contact_url']:
            lines.append(f"Profile: {draft['contact_url']}\n")
        lines.append(f"Characters: {len(draft['message'])}\n\n")
        lines.append(f"{draft['message']}\n\n")
        lines.append('-' * 40 + '\n\n')

    with open(filepath, 'w', encoding='utf-8') as f:
        f.writelines(lines)

    print(f"  Saved: {filepath}")
    return filepath, len(drafts)


def _update_log(company, roles, action):
    """Append to outreach_log.txt with timestamp."""
    log_path = os.path.join(OUTPUT_DIR, 'outreach_log.txt')
    roles_str = ", ".join(roles)
    with open(log_path, 'a', encoding='utf-8') as f:
        f.write(f"{datetime.now().strftime('%Y-%m-%d %H:%M')} | {action} | {company} | Roles: {roles_str}\n")


def run_outreach():
    """Main outreach function: reads tracker, skips existing, generates new drafts."""
    print("\n--- Outreach Drafter ---")
    os.makedirs(OUTPUT_DIR, exist_ok=True)

    applied = read_applied_companies()
    if not applied:
        print("  No applied companies found. Nothing to draft.")
        return

    generated = 0
    skipped = 0
    total_drafts = 0

    for company, data in applied.items():
        roles = data['roles']
        contacts = data['contacts']

        if not contacts:
            print(f"  Skipped {company}: no HR contacts")
            skipped += 1
            continue

        if has_existing_draft(company, roles):
            print(f"  Skipped {company}: drafts up to date")
            _update_log(company, roles, "SKIPPED (up to date)")
            skipped += 1
            continue

        filepath, draft_count = generate_outreach(company, roles, contacts)
        _update_log(company, roles, "GENERATED")
        generated += 1
        total_drafts += draft_count

    print(f"\n  Outreach complete: {generated} companies drafted ({total_drafts} messages), {skipped} skipped")


def main():
    """Entry point for standalone execution."""
    run_outreach()


if __name__ == "__main__":
    main()
