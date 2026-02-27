"""
mark_rejected.py — mark a company as Rejected in List.xlsx

Sets column D to "Rejected" and applies strikethrough font across columns A-F.
Append-only safe: only touches the matched company row(s).

Usage:
  python .claude/skills/mark-rejected/scripts/mark_rejected.py "Company Name"
"""
import sys
import os
import shutil
import tempfile
import openpyxl
from openpyxl.styles import Font

sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..', '..', '..', '..'))
try:
    from config import TRACKER_FILE
except ImportError:
    print("ERROR: config.py not found. Copy config.template.py to config.py.")
    sys.exit(1)

STRIKETHROUGH_COLS = 6  # A through F


def apply_strikethrough(ws, row_idx):
    for col in range(1, STRIKETHROUGH_COLS + 1):
        cell = ws.cell(row=row_idx, column=col)
        existing = cell.font
        cell.font = Font(
            name=existing.name,
            size=existing.size,
            bold=existing.bold,
            italic=existing.italic,
            color=existing.color,
            strike=True,
        )


def mark_rejected(company_name):
    search = company_name.strip().lower()
    temp_file = None
    try:
        try:
            wb = openpyxl.load_workbook(TRACKER_FILE)
        except PermissionError:
            print("  File locked — using temp copy...")
            fd, temp_file = tempfile.mkstemp(suffix='.xlsx')
            os.close(fd)
            shutil.copy2(TRACKER_FILE, temp_file)
            wb = openpyxl.load_workbook(temp_file)

        ws = wb.active
        matched = []

        for row in ws.iter_rows(min_row=2):
            cell_a = row[0]
            if cell_a.value and search in str(cell_a.value).strip().lower():
                row_idx = cell_a.row
                # Set status column (D = col 4)
                ws.cell(row=row_idx, column=4).value = "Rejected"
                apply_strikethrough(ws, row_idx)
                matched.append((row_idx, str(cell_a.value).strip()))

        if not matched:
            print(f"No rows found matching: '{company_name}'")
            return

        wb.save(TRACKER_FILE)
        for idx, name in matched:
            print(f"  Row {idx}: '{name}' -> Rejected + strikethrough applied")
        print(f"Done. {len(matched)} row(s) updated.")

    finally:
        if temp_file and os.path.exists(temp_file):
            os.remove(temp_file)


if __name__ == '__main__':
    if len(sys.argv) < 2:
        print("Usage: python mark_rejected.py \"Company Name\"")
        sys.exit(1)
    mark_rejected(sys.argv[1])
