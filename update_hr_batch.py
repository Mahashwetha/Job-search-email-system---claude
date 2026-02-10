"""Batch update of HR contacts - restore originals + append new from expanded query."""
import openpyxl
import re
from openpyxl.styles import Font, Alignment

TRACKER_FILE = r'C:\Users\mahas\OneDrive\Desktop\Applications\List.xlsx'

# TRUE ORIGINAL contacts from update_hr_contacts.py (the first script ever run)
TRUE_ORIGINAL = {
    "Pigment": [
        ("Nandi Toroosian", "https://www.linkedin.com/in/nanditoroosian/"),
        ("Julien Benavente", "https://fr.linkedin.com/in/julien-benavente-106a5637"),
    ],
    "Keyrock": [
        ("Nabanita Banerjee", "https://www.linkedin.com/in/nabanita-banerjee/"),
    ],
    "Euronext": [
        ("Mendy Abessira", "https://www.linkedin.com/in/mendy-abessira/"),
        ("Nathalie Rooney", "https://www.linkedin.com/in/nathalie-rooney/"),
    ],
    "Datadog": [
        ("Yuan HU", "https://fr.linkedin.com/in/yuanhu6/en"),
        ("Adeline Medin", "https://www.linkedin.com/in/adeline-m%C3%A9din-32321773/"),
    ],
    "Netflix": [
        ("Chelsea Chiappe", "https://www.linkedin.com/in/chelseachiappe/"),
        ("Nathan Richards", "https://www.linkedin.com/in/nathanmrichards/"),
    ],
    "Mirakl": [
        ("Sihem Y.", "https://fr.linkedin.com/in/sihemyahla/en"),
        ("Alexandre Martin", "https://www.linkedin.com/in/alexandre-martin-aix/"),
    ],
    "Station F": [
        ("(Contact via website)", "https://stationf.co/"),
    ],
    "Alan": [
        ("Jean-Charles Samuelian-Werve", "https://fr.linkedin.com/in/jcsamuelian"),
    ],
    "Stripe": [
        ("Reema Polignano", "https://www.linkedin.com/in/parikhrs/"),
        ("Emilie Schwartz", "https://www.linkedin.com/in/emilie-schwartz-16057b69/"),
    ],
    "Sanofi": [
        ("Cedric ELLORE", "https://fr.linkedin.com/in/c%C3%A9dric-ellore-dric971"),
        ("Tarana Verma", "https://www.linkedin.com/in/taranaverma/"),
    ],
    "Doctolib": [
        ("Salome Amsler", "https://fr.linkedin.com/in/salom%C3%A9-amsler-b3ab8655"),
        ("Sylvie Jordan", "https://www.linkedin.com/in/sylvie-jordan-peopleperson/"),
        ("Loic Le Bescond", "https://www.linkedin.com/in/lo%C3%AFc-le-bescond-b92aa097/"),
    ],
    "Ledger": [
        ("Paul Baudouin", "https://fr.linkedin.com/in/paulbaudouin89"),
        ("Sophie-Gaelle Waffo Wokam", "https://www.linkedin.com/in/sophie-ga%C3%ABlle-waffowokam/"),
    ],
    "Murex": [
        ("Severine Beaupere", "https://www.linkedin.com/in/s%C3%A9verine-beaup%C3%A8re-b574037/"),
        ("Mohamad Najda", "https://www.linkedin.com/in/mohamad-najda/"),
    ],
    "Finastra": [
        ("Belen Quiroga", "https://www.linkedin.com/in/belenquiroga/"),
        ("Nina van der Beugel", "https://www.linkedin.com/in/nina-van-der-beugel-99a41732/"),
    ],
    "JP Morgan": [
        ("Div Grover", "https://www.linkedin.com/in/div-grover-50a1277/"),
        ("Timothy F. Hunt", "https://www.linkedin.com/in/timothyhuntjpmc/"),
    ],
    "Morgan Stanley": [
        ("Steph Ahrens", "https://uk.linkedin.com/in/steph-ahrens-7b97301"),
        ("Jennifer Daley", "https://www.linkedin.com/in/jennifer-daley-21aa04b/"),
    ],
    "Factset": [
        ("Brianne Markey", "https://www.linkedin.com/in/brianne-markey-b991726/"),
        ("Hadrien P.", "https://www.linkedin.com/in/hadrien-proust/"),
    ],
    "LSEG": [
        ("Rosie Bartholomew", "https://www.linkedin.com/in/rosie-bartholomew-3899a9118/"),
        ("Lei Cao", "https://www.linkedin.com/in/lei-cao-6346a9b/"),
    ],
    "Google": [
        ("Anais Laine", "https://fr.linkedin.com/in/ana%C3%AFs-lain%C3%A9-1871ab35"),
        ("Penelope Walker", "https://fr.linkedin.com/in/penelope-walker-873013200"),
        ("Laura Dupas", "https://www.linkedin.com/in/laura-dupas-4940b5161/"),
    ],
    "Cisco": [
        ("Randy Norman", "https://www.linkedin.com/in/recruiterrandy/"),
        ("Kelly Olson", "https://www.linkedin.com/in/kellycallahanolson/"),
    ],
    "Visa": [
        ("Maurice Garner", "https://www.linkedin.com/in/maurice-garner-b167162/"),
        ("Quran Randall", "https://www.linkedin.com/in/quran-randall-575065104/"),
    ],
    "The fork/trip advisor": [
        ("Mathilde Blin", "https://www.linkedin.com/in/mathilde-blin"),
    ],
    "Sage": [
        ("Mike Durkin", "https://www.linkedin.com/in/themikedurkin/"),
        ("Emily Botero", "https://www.linkedin.com/in/emily-pearce-botero-859a36127/"),
    ],
    "BPCE SA": [
        ("Taoufik Mechmoum", "https://www.linkedin.com/in/taoufik-mechmoum-949574100/"),
        ("Sandrine Sullivan", "https://www.linkedin.com/in/sandrine-sullivan-2a83683b/"),
    ],
    "Ripple": [
        ("Matthew Welch", "https://www.linkedin.com/in/welchmatthew/"),
        ("James Warner", "https://www.linkedin.com/in/james-warner-66380753/"),
    ],
    "BforBank": [
        ("Aurelie Guillou-Duhamel", "https://www.linkedin.com/in/aur%C3%A9lie-guillou-duhamel-b8380110/"),
        ("Melissa Tanguy", "https://www.linkedin.com/in/m%C3%A9lissa-tanguy/"),
    ],
    "Akur8": [
        ("Morgan Le Louer", "https://www.linkedin.com/in/morganlelouer/"),
        ("Chiara Montalbano", "https://www.linkedin.com/in/chiara-montalbano-1b8967152/"),
    ],
    "dfns": [
        ("Christopher Grilhault des Fontaines", "https://www.linkedin.com/in/christopherdesfontaines/"),
        ("Joy Sparrowhawk", "https://www.linkedin.com/in/joy-sparrowhawk/"),
    ],
    "Welcome to jungle by station F": [
        ("(Search WelcomeToTheJungle)", "https://www.welcometothejungle.com/"),
    ],
    "Revolut": [
        ("Alejandro Lopez Arnau", "https://www.linkedin.com/in/alejandro-talent-specialist/"),
        ("Nadin El-Rifai", "https://www.linkedin.com/in/nadin-el-rifai-6a804314a/"),
        ("Mihai Pop", "https://www.linkedin.com/in/mihai-pop-a70063208/"),
    ],
    "Pelico": [
        ("Etienne Vial", "https://www.linkedin.com/in/etienne-vial-29438/"),
        ("Axel Branger", "https://www.linkedin.com/in/axel-branger/"),
    ],
    "BNP Paribas": [
        ("Makeda Tchikaty", "https://www.linkedin.com/in/mak%C3%A9da-tchikaty-b42233b4/"),
        ("Sumayya M", "https://www.linkedin.com/in/sumayya-m-80baa5184/"),
    ],
    "Nasdaq": [
        ("Aurelie Behar", None),  # User's manual edit replacing India-based contacts
    ],
    "Trade Republic": [
        ("Palak Mehta", "https://www.linkedin.com/in/palak-mehta-86837bb7/"),
        ("Marina Fignani", "https://www.linkedin.com/in/marina-fignani/"),
    ],
    "Payflows": [
        ("Tidia O.", "https://www.linkedin.com/in/tidia/"),
    ],
    "qonto": [
        ("Diane Levron", "https://www.linkedin.com/in/diane-levron-monteux-7b899310b/"),
        ("Chi Nguyen", "https://www.linkedin.com/in/chi-nguyen-mai/"),
        ("Alexandra Lyraud", "https://fr.linkedin.com/in/alexandra-lyraud-b9b450194/"),
    ],
    "Broadridge": [
        ("alice", None),  # User's manual entry
    ],
    "Criteo": [
        ("Celia Navarro", "https://fr.linkedin.com/in/cenavarro"),
        ("Leonore Lahalle", "https://fr.linkedin.com/in/l%C3%A9onore-lahalle-a8212a29"),
        ("Cristina Craciun", "https://www.linkedin.com/in/cristinaacraciun/"),
    ],
    # User manual edits from email (overriding script values)
    "Keyrock_manual": [
        ("chloe hudson", None),
        ("ANOTHONY S", None),
    ],
    "Euronext_manual_extra": [
        ("housnate", None),  # Added by user before the script contacts
    ],
}

# NEW contacts from expanded query (Feb 2026) - only truly new names
# These get APPENDED, never replace
NEW_EXTRA_CONTACTS = {
    "Pigment": [
        ("Julia Gaulter", "https://www.linkedin.com/in/juliagaulter/"),
    ],
    "Keyrock": [
        ("Rebecca Neal", "https://www.linkedin.com/in/rebecca-neal-93301398/"),
    ],
    "Euronext": [
        ("Mette Wassmann", "https://www.linkedin.com/in/mettewassmann/"),
    ],
    "Datadog": [
        ("Marianne Fournier", "https://www.linkedin.com/in/mariannefournier/"),
        ("Alison Vannier", "https://www.linkedin.com/in/alisonvannier/"),
    ],
    "Netflix": [
        ("Marta Gutierrez", "https://mx.linkedin.com/in/martagucla"),
        ("Joe Wright", "https://www.linkedin.com/in/joe-wright-41b50092/"),
    ],
    "Sanofi": [
        ("Urszula Pazur", "https://www.linkedin.com/in/urszulapazur/"),
    ],
    "Ledger": [
        ("Laetitia Kunz", "https://fr.linkedin.com/in/laetitiakunz/en"),
    ],
    "JP Morgan": [
        ("Helena Sharpe", "https://uk.linkedin.com/in/helenasharpe"),
        ("Karen Brennock", "https://www.linkedin.com/in/karenbrennock/"),
    ],
    "Morgan Stanley": [
        ("Giulia Moscatellini", "https://www.linkedin.com/in/giulia-moscatellini-20146b1a0/"),
    ],
    "Factset": [
        ("Alexandra Dixon", "https://www.linkedin.com/in/alexandralanyi/"),
    ],
    "LSEG": [
        ("Jo Clark", "https://www.linkedin.com/in/joanneclark2/"),
    ],
    "Cisco": [
        ("Maria Da Costa", "https://fr.linkedin.com/in/mariadacostacisco"),
    ],
    "Visa": [
        ("Andrea Sola", "https://www.linkedin.com/in/andr%C3%A9asola/"),
        ("Scott Newman", "https://www.linkedin.com/in/scott-newman-35a953a/"),
    ],
    "Sage": [
        ("Christine Laux", "https://www.linkedin.com/in/christinelaux/"),
    ],
    "BPCE SA": [
        ("Aude Nuytten", "https://www.linkedin.com/in/aude-nuytten%E2%9C%A8-24513a122/"),
        ("Annick Leveziel", "https://be.linkedin.com/in/annick-leveziel-b4a80110/"),
    ],
    "Ripple": [
        ("Jasmine Cooper", "https://www.linkedin.com/in/jmcooper1/"),
        ("Joakim Danielsson", "https://www.linkedin.com/in/joakimdanielsson/"),
    ],
    "BforBank": [
        ("Aline Fale", "https://www.linkedin.com/in/aline-fal%C3%A9-60a26a108/"),
        ("Claire Di Filippo", "https://www.linkedin.com/in/claire-di-filippo/"),
    ],
    "Revolut": [
        ("Manuela Gomez", "https://www.linkedin.com/in/manuela-g%C3%B3mez-s%C3%A1nchez-arambur%C3%BA-767b10169/"),
        ("Sofia G.", "https://www.linkedin.com/in/sofia-gil/"),
    ],
    "Criteo": [
        ("Pauline Billotte", "https://fr.linkedin.com/in/billottepauline/en"),
    ],
    "BNP Paribas": [
        ("Helene Trouillet", "https://www.linkedin.com/in/h%C3%A9l%C3%A8ne-trouillet-bnpp-2015/"),
    ],
    "Nasdaq": [
        ("Greta P.", "https://es.linkedin.com/in/gretapaulauskaite"),
        ("Sanna Hedlund", "https://www.linkedin.com/in/sanna-hedlund/"),
    ],
    "Broadridge": [
        ("Ciara Salina", "https://www.linkedin.com/in/ciara-salina-b0934b69/"),
        ("Patricia Priola", "https://www.linkedin.com/in/patricia-priola/"),
    ],
    "Nutanix": [
        ("Kevin Munoz", "https://www.linkedin.com/in/kevinmunozsantos/"),
        ("Nathalie Gryba", "https://www.linkedin.com/in/nathalieravenstein/"),
    ],
}


def find_match(company_name, contacts_dict):
    if company_name in contacts_dict:
        return contacts_dict[company_name]
    for key in contacts_dict:
        if '_manual' in key:
            continue  # Skip manual override keys in generic matching
        if (key.lower() == company_name.lower()
                or key.lower() in company_name.lower()
                or company_name.lower() in key.lower()):
            return contacts_dict[key]
    return None


def build_formula(contacts):
    """Build Excel formula from list of (name, url_or_None) tuples."""
    parts = []
    for name, url in contacts:
        if url:
            parts.append(f'HYPERLINK("{url}","{name}")')
        else:
            parts.append(f'"{name}"')
    if len(parts) == 1:
        if contacts[0][1]:
            return f'=HYPERLINK("{contacts[0][1]}","{contacts[0][0]}")'
        else:
            return contacts[0][0]  # Plain text, no formula
    return "=" + " & CHAR(10) & ".join(parts)


def write_contacts(ws, row_idx, contacts, link_font):
    """Write contacts to HR cell."""
    hr_cell = ws.cell(row=row_idx, column=5)
    hr_cell.number_format = 'General'  # Ensure Excel evaluates formulas
    hr_cell.value = build_formula(contacts)
    hr_cell.font = link_font
    hr_cell.alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[row_idx].height = max(30, len(contacts) * 15)


def extract_existing_names(cell_value):
    """Extract contact names from cell."""
    if not cell_value:
        return set()
    val = str(cell_value)
    names = re.findall(r'HYPERLINK\([^,]+,"([^"]+)"\)', val)
    # Also check for plain text names in formula like "alice"
    plain = re.findall(r'(?<!=)"([^"]+)"(?!\))', val)
    names.extend(plain)
    if not names and val and not val.startswith('='):
        names = [val.strip()]
    return set(n.lower() for n in names)


def extract_existing_contacts(cell_value):
    """Extract (name, url) tuples from cell."""
    if not cell_value:
        return []
    val = str(cell_value)
    pairs = re.findall(r'HYPERLINK\("([^"]+)","([^"]+)"\)', val)
    contacts = [(name, url) for url, name in pairs]
    # Also extract plain text names
    plain = re.findall(r'(?<!=)"([^"]+)"(?!\))', val)
    for p in plain:
        if not any(p.lower() == n.lower() for n, _ in contacts):
            contacts.append((p, None))
    if not contacts and val and not val.startswith('='):
        contacts.append((val.strip(), None))
    return contacts


def main():
    wb = openpyxl.load_workbook(TRACKER_FILE, data_only=False)
    ws = wb.active
    link_font = Font(color="0563C1", underline="single", size=10)

    updated = 0
    for row_idx in range(2, ws.max_row + 1):
        company_cell = ws.cell(row=row_idx, column=1)
        company = str(company_cell.value).strip() if company_cell.value else ""
        if not company or "Program/Product" in company:
            continue

        # Get original contacts
        if company.lower() == "keyrock":
            base = list(TRUE_ORIGINAL.get("Keyrock_manual", []))
        elif company.lower() == "euronext":
            extra = TRUE_ORIGINAL.get("Euronext_manual_extra", [])
            base = list(extra) + list(TRUE_ORIGINAL.get("Euronext", []))
        else:
            base = list(find_match(company, TRUE_ORIGINAL) or [])

        if not base:
            continue

        # Get new extra contacts and filter duplicates
        extras = find_match(company, NEW_EXTRA_CONTACTS) or []
        existing_names = set(n.lower() for n, _ in base)
        new_to_add = [(n, u) for n, u in extras if n.lower() not in existing_names]

        # Combine: original + new extras, write in one shot
        all_contacts = base + new_to_add
        write_contacts(ws, row_idx, all_contacts, link_font)
        updated += 1

        if new_to_add:
            print(f"  Row {row_idx}: {company} = {', '.join(n for n, _ in base)} += {', '.join(n for n, _ in new_to_add)}")

    wb.save(TRACKER_FILE)
    wb.close()
    print(f"\nUpdated {updated} rows total")


if __name__ == "__main__":
    main()
